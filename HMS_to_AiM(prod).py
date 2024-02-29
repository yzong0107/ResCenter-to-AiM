__author__ = "Tim Zong(yzong@ualberta.ca)"
import time
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support import expected_conditions
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.desired_capabilities import DesiredCapabilities
import getpass
import traceback
import openpyxl
from selenium.webdriver.support.ui import Select
from datetime import datetime
from selenium.webdriver.edge.service import Service
from webdriver_manager.microsoft import EdgeChromiumDriverManager
import os

class AiM():
    def setup_method(self):
        options = webdriver.EdgeOptions()
        options.add_argument("-inprivate")
        self.driver = webdriver.Edge(service=Service(EdgeChromiumDriverManager().install()),options=options)
        self.driver.maximize_window()
        self.vars = {}

    def teardown_method(self):
        self.driver.quit()

    def login(self):
        self.driver.get("https://www.aimprod.ualberta.ca/fmax/login")
        self.driver.maximize_window()
        # self.driver.set_window_size(1900, 1020)
        username = input('Enter your username (AiM): ')
        password = getpass.getpass('Enter your password (AiM): ')
        self.driver.find_element(By.ID, "username").send_keys(username)
        self.driver.find_element(By.ID, "password").send_keys(password)
        self.driver.find_element(By.ID, "login").click()

    def customer_request(self,description,reference,location,property="51000"):
        self.driver.find_element(By.ID, "mainForm:menuListMain:CUSTSERVICE").click()
        self.driver.find_element(By.ID, "mainForm:menuListMain:new_CRQ_VIEW").click()
        aim_des = location + " - " + description
        time.sleep(1)
        self.driver.find_element(By.ID, "mainForm:CRQ_EDIT_content:ae_p_req_e_description").send_keys(aim_des)
        time.sleep(0.5)
        self.driver.find_element(By.ID, "mainForm:CRQ_EDIT_content:locZoomType1:locZoomType1-2").send_keys(property)
        self.driver.find_element(By.CSS_SELECTOR, "#mainForm\\3A CRQ_EDIT_content\\3AlocZoomType1\\3AlocZoomType1-2_button > .halflings").click()
        time.sleep(1)
        self.driver.find_element(By.ID, "mainForm:CRQ_EDIT_content:referenceNoValueType1").send_keys(reference)
        self.driver.find_element(By.ID,"mainForm:CRQ_EDIT_content:connamevalueType1").send_keys("CS Space IWC")
        self.driver.find_element(By.ID,"mainForm:CRQ_EDIT_content:conphValueType1").send_keys("780-492-6045")
        self.driver.find_element(By.ID,"mainForm:CRQ_EDIT_content:conmcValueType1").send_keys("asiwc@ualberta.ca")

        if len(aim_des)>255: #extends word limit in normal description area
            self.driver.find_element(By.ID,"mainForm:sideButtonPanel:moreMenu_0").click()
            self.driver.find_element(By.ID,"mainForm:ae_p_req_e_long_desc").send_keys(aim_des)
            self.driver.find_element(By.ID,"mainForm:buttonPanel:done").click()
        time.sleep(1)
        self.driver.find_element(By.ID, "mainForm:buttonPanel:save").click()
        aim_CR = self.driver.find_element(By.ID,"mainForm:CRQ_VIEW_content:ae_p_req_e_doc_no").text
        self.driver.find_element(By.ID,"mainForm:headerInclude:aimTitle1").click()
        return aim_CR



class ResCenter():
    def setup_method(self):
        options = webdriver.EdgeOptions()
        options.add_argument("-inprivate")
        self.driver = webdriver.Edge(service=Service(EdgeChromiumDriverManager().install()),options=options)
        self.driver.maximize_window()
        self.vars = {}
        self.time_out_sec = 10

    def teardown_method(self):
        self.driver.quit()

    def login(self):
        self.driver.get("https://rezsrv.ancillary.ualberta.ca/ResCenter/Login/")
        self.driver.maximize_window()
        # self.driver.set_window_size(1900, 1020)
        username = input('Enter your username (ResCenter): ')
        password = getpass.getpass('Enter your password (ResCenter): ')
        self.driver.find_element(By.ID, "ctl00_mainContent_txtUserName_cbTextBox").send_keys(username)
        self.driver.find_element(By.ID, "ctl00_mainContent_txtPassword_cbTextBox").send_keys(password)
        self.driver.find_element(By.CSS_SELECTOR, "#ctl00_mainContent_btnLogin_CBORDLinkButton > .btn-text").click()

    def search(self):
        # self.driver.find_element(By.CSS_SELECTOR, ".fa-bars").click()
        try: 
            WebDriverWait(self.driver, self.time_out_sec).until(EC.presence_of_element_located((By.CSS_SELECTOR, ".fa-close")))
            self.driver.find_element(By.CSS_SELECTOR, ".fa-close").click()
            time.sleep(3)
        except:
            pass
        WebDriverWait(self.driver, self.time_out_sec).until(EC.presence_of_element_located((By.ID, "FacilitiesModule")))
        self.driver.find_element(By.ID, "FacilitiesModule").click()
        time.sleep(3)
        WebDriverWait(self.driver, self.time_out_sec).until(EC.presence_of_element_located((By.LINK_TEXT, "Maintenance")))
        self.driver.find_element(By.LINK_TEXT, "Maintenance").click()
        WebDriverWait(self.driver, self.time_out_sec).until(EC.presence_of_element_located((By.LINK_TEXT, "Work Orders")))
        self.driver.find_element(By.LINK_TEXT, "Work Orders").click()
        self.driver.find_element(By.ID, "ctl00_mainContent_ddWOStatus").click()
        dropdown = self.driver.find_element(By.ID, "ctl00_mainContent_ddWOStatus")
        dropdown.find_element(By.XPATH, "//option[. = 'NEW']").click()
        self.driver.find_element(By.ID, "ctl00_mainContent_ddWOStatus").click()
        time.sleep(3)
        # """test for handling last record, and end the program"""
        # time.sleep(0.5)
        # self.driver.find_element(By.XPATH, "//select[@name='ctl00$mainContent$ddWOType']/option[text()='UCR']").click()
        # """    """
        self.driver.find_element(By.CSS_SELECTOR, "#ctl00_mainContent_btnRunSearch_CBORDLinkButton > .btn-text").click()

    def top_record(self):
        try:
            WebDriverWait(self.driver, self.time_out_sec).until(EC.presence_of_element_located((By.ID, "ctl00_mainContent_radgridWorkOrders_ctl00_ctl04_btnSelect_CBORDLinkButton")))
            self.driver.find_element(By.ID, "ctl00_mainContent_radgridWorkOrders_ctl00_ctl04_btnSelect_CBORDLinkButton").click()
        except: #Timeout exception
            return (None,None,None,None,None)
        WebDriverWait(self.driver, self.time_out_sec).until(EC.presence_of_element_located((By.CSS_SELECTOR, "#ctl00_mainContent_txtDescription_ReadOnlyBox pre")))
        time.sleep(1)
        description_res = self.driver.find_element(By.CSS_SELECTOR, "#ctl00_mainContent_txtDescription_ReadOnlyBox pre").text
        WO_res = self.driver.find_element(By.ID, "ctl00_mainContent_txtWOIDNum_cbTextBox").get_attribute("value")
        location_res = self.driver.find_element(By.ID, "ctl00_mainContent_FacilityLookup_txtFacilityNameSearch").get_attribute("value")
        options = Select(self.driver.find_element(By.ID, "ctl00_mainContent_ddWOType"))
        WO_type = options.first_selected_option.text
        email_to_notify = self.driver.find_element(By.ID,"ctl00_mainContent_txtRqstrEmail_cbTextBox").get_attribute("value")
        staff_email_to_notify = self.driver.find_element(By.ID,"ctl00_mainContent_txtNotifyEmail_cbTextBox").get_attribute("value")
        if len(staff_email_to_notify)!=0:
            email_to_notify = ","+email_to_notify
        return (WO_res,description_res,location_res,WO_type,email_to_notify)

    def edit(self,aim_cr,email):
        WebDriverWait(self.driver, self.time_out_sec).until(EC.presence_of_element_located((By.ID,"ctl00_mainContent_btnTopEdit_CBORDLinkButton")))
        self.driver.find_element(By.ID,"ctl00_mainContent_btnTopEdit_CBORDLinkButton").click()
        time.sleep(2) # explicitly wait for 2 seconds before the page is updated

        """March 09, 2020 copy paste email address"""
        self.driver.find_element(By.ID, "ctl00_mainContent_txtNotifyEmail_cbTextBox").send_keys(email)
        """"End"""

        if aim_cr is not None:
            
            options = Select(self.driver.find_element(By.ID, "ctl00_mainContent_ddWOType"))
            current_type = options.first_selected_option.text
            if current_type!="Pest Control":
                self.driver.find_element("xpath","//select[@name='ctl00$mainContent$ddWOType']/option[text()='General']").click() #202402: if it's pest control, leave it
                canned_response = "Your request has been Assigned to Maintenance Staff (or another UofA Staff) who will visit your unit / room to respond to the issue(s) reported in order of all the priorities in Residence. This email serves as a Notice of Entry and is effective from 10 business days of being \"Assigned\" in our system. Urgent issues will be expedited. Thank you for your patience as Staff work diligently to address all issues and concerns reported.\n"
            else:
                canned_response="Your request has been Assigned to a Pest Control Technician who will visit your unit / room to respond to the issue(s) reported in order of all the priorities in Residence. This email serves as a Notice of Entry and is effective from 10 business days of being \"Assigned\" in our system. Urgent issues will be expedited. Thank you for your patience as Staff work diligently to address all issues and concerns reported.\n"
            
            actions = ActionChains(self.driver)
            actions.move_to_element(self.driver.find_element(By.CSS_SELECTOR, "#ctl00_mainContent_txtDescription_FancyTextBoxTextArea"))
            actions.click()
            actions.key_down(Keys.CONTROL)
            actions.send_keys(Keys.HOME)
            actions.key_up(Keys.CONTROL)
            actions.send_keys(aim_cr+": "+canned_response) #202402: remove "AiM CR"
            actions.perform()
            self.driver.find_element("xpath","//select[@name='ctl00$mainContent$ddWOStatus']/option[text()='Assigned']").click()
            time.sleep(0.5)
        else:
            self.driver.find_element("xpath","//select[@name='ctl00$mainContent$ddWOStatus']/option[text()='Pending ASIWC Approval']").click()

        time.sleep(0.5)
        self.driver.find_element("xpath","//select[@name='ctl00$mainContent$ddWOPriority']/option[text()='Normal']").click()
        time.sleep(0.5)

        self.driver.find_element(By.ID, "ctl00_mainContent_btnTopSave_CBORDLinkButton").click()
        WebDriverWait(self.driver, self.time_out_sec).until(EC.presence_of_element_located((By.CSS_SELECTOR, " .alert")))
        # self.driver.find_element(By.ID, "ctl00_mainContent_btnTopClose_CBORDLinkButton").click()
        pop_message = self.driver.find_element(By.CSS_SELECTOR,".alert > span").text
        if pop_message=="Work Order successfully updated":
            self.driver.find_element(By.ID,"ctl00_mainContent_btnTopClose_CBORDLinkButton").click()
            return True,None
        else: # location cell missing
            error = self.driver.find_element(By.CSS_SELECTOR,"ul:nth-child(3)").text
            self.driver.find_element(By.CSS_SELECTOR,".fa-close").click()

            self.driver.find_element(By.ID,"ctl00_mainContent_FacilityLookup_txtFacilityNameSearch").send_keys("ASIWC Office") # hard coded to a fixed value
            self.driver.find_element("xpath","//select[@name='ctl00$mainContent$ddWOStatus']/option[text()='Pending ASIWC Approval']").click()  # hard coded to a fixed value
            options = Select(self.driver.find_element(By.ID, "ctl00_mainContent_ddWOType"))
            current_type = options.first_selected_option.text
            if current_type!="Pest Control":
                self.driver.find_element("xpath","//select[@name='ctl00$mainContent$ddWOType']/option[text()='General']").click() #202402: if it's pest control, leave it
            self.driver.find_element("xpath","//select[@name='ctl00$mainContent$ddWOPriority']/option[text()='Normal']").click()

            self.driver.find_element(By.ID,"ctl00_mainContent_FacilityLookup_btnSearch_CBORDLinkButton").click()
            time.sleep(1)
            self.driver.find_element(By.ID, "ctl00_mainContent_btnTopSave_CBORDLinkButton").click()
            WebDriverWait(self.driver, self.time_out_sec).until(EC.presence_of_element_located((By.CSS_SELECTOR, " .alert")))
            self.driver.find_element(By.ID,"ctl00_mainContent_btnTopClose_CBORDLinkButton").click()
            return False,error

if __name__ == '__main__':
    start_time = time.time()

    aim_window = AiM()
    res_window = ResCenter()

    aim_window.setup_method()
    res_window.setup_method()

    aim_window.login()
    res_window.login()

    res_window.search()

    count = 0
    try:
        dir_path = os.path.dirname(os.path.realpath(__file__))
        wb = openpyxl.load_workbook(dir_path+"/Logs.xlsx")
        ws = wb.worksheets[0]
        id = ws.cell(row=ws.max_row,column=1).value # get the id of last row
        if id=="ID":
            id = 0 # initial id
        while True:
            id += 1
            res_Wo,res_des,res_loc,WO_type,email_to_notify = res_window.top_record() # read top record in ResCenter

            # 20240223: No need to differentiate Pest Control or Contractor from other WO types
            # if WO_type=="Pest Control" or WO_type=="Contractor":
            #     saved, error_message = res_window.edit(aim_cr=None,email=email_to_notify)  # update in ResCenter
            #     if saved:
            #         new_row = [id, res_Wo, None, "Not Processed",datetime.now(),"Need further review, as WO type is {}.".format(WO_type)]
            #         print ("ResCenter WO# {0} has NOT been logged into AiM!".format(res_Wo))
            #     else:
            #         extra_notes = "Need further review, as WO type is {}.\n".format(WO_type)
            #         new_row = [id, res_Wo, None, "Not Processed", datetime.now(), extra_notes+error_message+" Default location applied."]
            #         print ("ResCenter WO# {0} has NOT been logged into AiM!".format(res_Wo))
            #     ws.append(new_row)
            #     wb.save("Logs.xlsx")
            if res_Wo is not None: # processed all data
                aim_CR = aim_window.customer_request(res_des,res_Wo,res_loc) # Log AiM Customer Request
                saved,error_message= res_window.edit(aim_CR,email_to_notify) # update in ResCenter
                if saved:
                    new_row = [id, res_Wo, aim_CR, "Processed",datetime.now(),""]
                    print ("ResCenter WO# {0} has been logged into AiM!".format(res_Wo))
                else:
                    new_row = [id, res_Wo, aim_CR, "Processed(with error)", datetime.now(), error_message+" Default location applied."]
                    print ("ResCenter WO# {0} has been logged into AiM! {1}".format(res_Wo,error_message+" Default location applied."))
                ws.append(new_row)
                wb.save(dir_path+"/Logs.xlsx")
            else:
                break
            count += 1

    except:
        # if there is any other errors, stops
        print(traceback.format_exc())

    time_taken = time.time()-start_time
    print ("")
    print ("***************************************")
    print("Finished! {} of records are processed!".format(count))
    print ("Time taken: {:.2f}s ({:.2f}min)".format(time_taken,time_taken/60.))
    print("***************************************")
